from client_manager import ClientManager
from appointment_manager import AppointmentManager
from employee_manager import EmployeeManager
from tkinter import messagebox

import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter

# Add all of the necessary imports


def export_all_appointments_to_xlsx(
    appointment_manager, client_manager, destination_folder_path
):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    sheet = wb.active

    # Define the column headings for the spreadsheet
    headings = ["ID", "Client", "Email", "Phone", "Name", "Date", "Time (Minutes)"]

    # Write the column headings to the first row of the sheet
    for col_num, heading in enumerate(headings, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}1"]
        cell.value = heading
        cell.font = Font(bold=True)
        cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb="C6EFCE"))

    # Retrieve all appointments from the appointment manager
    appointments = appointment_manager.get_all_appointments()

    # Iterate over the appointments and populate the spreadsheet
    for row_num, appointment in enumerate(appointments, 2):
        # Retrieve the client information for the current appointment
        client = client_manager.get_client(appointment[4])

        # Write the appointment and client information to the corresponding cells in the sheet
        sheet[f"A{row_num}"] = appointment[0]  # ID
        sheet[f"B{row_num}"] = client[1] + " " + client[2]  # Client First and Last Name
        sheet[f"C{row_num}"] = client[4]  # Client Email
        sheet[f"D{row_num}"] = client[3]  # Client Phone
        sheet[f"E{row_num}"] = appointment[1]  # Appointment Name
        sheet[f"F{row_num}"] = appointment[2]  # Appointment Date
        sheet[f"G{row_num}"] = appointment[3]  # Appointment Duration

    # Set the filename and filepath for the exported spreadsheet
    filename = "appointments.xlsx"
    filepath = os.path.join(destination_folder_path, filename)

    # Save the workbook to the specified filepath
    wb.save(filepath)

    # Export message
    messagebox.showinfo(
        "Εξαγωγή σε Excel", "Η εξαγωγή του αρχείου ολοκληρώθηκε επιτυχώς"
    )

    # Return the filepath of the exported spreadsheet
    return filepath


def export_all_clients_to_xlsx(
    client_manager: ClientManager, destination_folder_path: str
):
    wb = Workbook()
    sheet = wb.active
    headings = ["ID", "First Name", "Last Name", "Phone", "Email"]

    # Write the column headings to the first row of the sheet
    for col_num, heading in enumerate(headings, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}1"]
        cell.value = heading
        cell.font = Font(bold=True)
        cell.fill = PatternFill(patternType="solid", fgColor=Color(rgb="C6EFCE"))

    # Retrieve all clients from the client manager
    clients = client_manager.get_all_clients()

    # Iterate over the clients and populate the spreadsheet
    for row_num, client in enumerate(clients, 2):
        # Write the client information to the corresponding cells in the sheet
        sheet[f"A{row_num}"] = client[0]
        sheet[f"B{row_num}"] = client[1]
        sheet[f"C{row_num}"] = client[2]
        sheet[f"D{row_num}"] = client[3]
        sheet[f"E{row_num}"] = client[4]

    # Set the filename and filepath for the exported spreadsheet
    filename = "clients.xlsx"
    filepath = os.path.join(destination_folder_path, filename)
    wb.save(filepath)

    # Export message
    messagebox.showinfo(
        "Εξαγωγή σε Excel", "Η εξαγωγή του αρχείου ολοκληρώθηκε επιτυχώς"
    )

    return filepath


def send_reminders_to_clients_at_date(
    appointment_manager, client_manager, date, employee
):
    # Retrieve the email and pass code of the employee
    email = employee[3]
    pass_code = employee[4]

    # Create an SMTP server object and connect to the Gmail SMTP server
    smtp_server = smtplib.SMTP("smtp.gmail.com", 587)
    smtp_server.starttls()

    # Login to the SMTP server using the employee's email and pass code
    smtp_server.login(email, pass_code)

    # Iterate over appointments on the specified date
    for appointment in appointment_manager.get_appointments_on_date(date):
        # Retrieve appointment information
        appointment_date = datetime.strptime(appointment[2], "%Y-%m-%d %H:%M:%S")

        # Set the locale to Greek for date formatting
        import locale

        locale.setlocale(locale.LC_TIME, "el_GR")

        # Format the appointment date and time
        date_time = appointment_date.strftime("%A, %d-%b-%Y, %H:%M:%S")

        # Retrieve client information for the appointment
        client = client_manager.get_client(appointment[4])

        # Compose the email message
        message = MIMEText(
            f"Αγαπητέ/ή {client[1]},\n\nΑυτή είναι μια φιλική υπενθύμιση ότι έχετε ραντεβού μαζί μας στις {date_time}.\nΠαρακαλούμε, ενημερώστε μας εκ των προτέρων εάν χρειαστεί να αναπρογραμματίσετε ή να ακυρώσετε το ραντεβού σας.\n\nΜε τους θερμότερους χαιρετισμούς,\n\nΗ ομάδα του SchedulEase."
        )

        # Set the email subject, sender, and recipient
        message["Subject"] = "Υπενθύμιση ραντεβού"
        message["From"] = email
        message["To"] = client[4]

        # Send the email to the client
        smtp_server.sendmail(email, client[4], message.as_string())

    # Disconnect from the SMTP server
    smtp_server.quit()


def get_stats_in_date_range(
    appointment_manager: AppointmentManager,
    employee_manager: EmployeeManager,
    start: datetime,
    end: datetime,
):
    # Retrieve all employees
    all_employees = employee_manager.get_all_employees()
    # List of tuples: [(id, name, email, pass_code)]

    # Create a list to store the statistics
    stats = []

    # Iterate over each employee
    for employee_id, employee_name, *_ in all_employees:
        # Construct the SQL query to count the number of appointments for the employee within the specified date range
        query = "SELECT COUNT(*) FROM appointments WHERE employee_id = ? AND date BETWEEN ? AND ?;"
        params = (employee_id, start, end)

        # Execute the query
        appointment_manager.cursor.execute(query, params)

        # Fetch the result of the query
        amount = appointment_manager.cursor.fetchone()[0]

        # Store the employee's statistics in the list
        stats.append({"name": employee_name, "amount": amount})

    # Return the statistics list
    return stats


def print_appointments_on_date(appointment_manager: AppointmentManager, date: datetime):
    # Get all appointments on the specified date
    appointments = appointment_manager.get_appointments_on_date(date)

    # Initialize the printer
    printer = None
    try:
        # Get the default printer
        printer = os.environ["PRINTER"]

        # Open the printer
        printer = open(printer, "w")

        # Print the header
        printer.write("Appointments on {}:\n".format(date))

        # Print each appointment
        for appointment in appointments:
            printer.write(
                "Name: {}\nDate: {}\nTime: {}\nDuration: {} minutes\nClient: {}\nEmployee: {}\n\n".format(
                    appointment["name"],
                    appointment["date"],
                    appointment["time"],
                    appointment["duration"],
                    appointment["client_id"],
                    appointment["employee_id"],
                )
            )

    finally:
        # Close the printer
        if printer is not None:
            printer.close()
